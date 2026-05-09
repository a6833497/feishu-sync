#!/usr/bin/env python3
"""
Lark 聊天文档全量导出脚本（替代 download-all-lark.js 的 Playwright UI 方式）

背景：download-all-lark.js 用 Playwright 点击「下载为 CSV」只拿到 ~10% 数据
（CSV 只 1048 distinct linky_id vs Lark 实际 27530+，且日期 cut-off 在 4-9）

方案：用 Lark Drive Export API（drive/v1/export_tasks）拿全量 xlsx：
1. wiki node token → obj_token (spreadsheet token)
2. export_tasks 创建任务 → 拿 ticket
3. 轮询任务完成 → 拿 file_token → 下载 xlsx
4. openpyxl parse xlsx 多 sheet → 写每个 csv 覆盖现有

2026-05-02 创建
"""
import json
import os
import sys
import time
import urllib.request
import urllib.parse
import csv
from datetime import datetime, timezone, timedelta

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(SCRIPT_DIR, "config.json")
EXPORT_DIR = os.path.join(SCRIPT_DIR, "lark-exports")
CST = timezone(timedelta(hours=8))

# 3 个聊天 wiki 文档（从 download-all-lark.js 抓出来的 url 末尾 token）
CHAT_DOCS = [
    {
        "name": "印尼聊天",
        "wiki_token": "DU3owAMqfidkyXklL6BjPgRApXf",
        "sheet_to_csv": {
            "日数据": "印尼聊天_日数据.csv",
            "周数据": "印尼聊天_周数据.csv",
            "日汇总": "印尼聊天_日汇总.csv",
            "周汇总": "印尼聊天_周汇总.csv",
            "印尼id": "印尼聊天_印尼id.csv",
            "官方数据": "印尼聊天_官方数据.csv",
            "裂变关系": "印尼聊天_裂变关系.csv",
            "周排行": "印尼聊天_周排行.csv",
            "裂变周数据": "印尼聊天_裂变周数据.csv",
            "带动消费对比": "印尼聊天_带动消费对比.csv",
        },
    },
    {
        "name": "巴西聊天",
        "wiki_token": "G29Cwov08iWq4Fk7kVPjW4K2pAe",
        "sheet_to_csv": {
            "日数据": "巴西聊天_日数据.csv",
            "周数据": "巴西聊天_周数据.csv",
            "日汇总": "巴西聊天_日汇总.csv",
            "周汇总": "巴西聊天_周汇总.csv",
            "巴西id": "巴西聊天_巴西id.csv",
            "官方数据": "巴西聊天_官方数据.csv",
            "裂变关系": "巴西聊天_裂变关系.csv",
            "裂变周数据": "巴西聊天_裂变周数据.csv",
            "周排名": "巴西聊天_周排名.csv",
        },
    },
    {
        "name": "西语聊天",
        "wiki_token": "Z50NwRS6cihkZJkdGwNj4wOzpse",
        "sheet_to_csv": {
            "日数据": "西语聊天_日数据.csv",
            "周数据": "西语聊天_周数据.csv",
            "日汇总": "西语聊天_日汇总.csv",
            "周汇总": "西语聊天_周汇总.csv",
            "西语id": "西语聊天_西语id.csv",
            "官方数据": "西语聊天_官方数据.csv",
            "裂变关系": "西语聊天_裂变关系.csv",
            "裂变周数据": "西语聊天_裂变周数据.csv",
            "裂变每周新增达标数": "西语聊天_裂变每周新增达标数.csv",
            "非直属转直属": "西语聊天_非直属转直属.csv",
            "国家数据分析": "西语聊天_国家数据分析.csv",
        },
    },
]


def log(msg):
    ts = datetime.now(CST).strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


def get_lark_token(config):
    lark = config["lark"]
    data = json.dumps({"app_id": lark["app_id"], "app_secret": lark["app_secret"]}).encode()
    req = urllib.request.Request(
        "https://open.larksuite.com/open-apis/auth/v3/tenant_access_token/internal",
        data=data,
        headers={"Content-Type": "application/json"},
    )
    resp = json.loads(urllib.request.urlopen(req).read())
    if resp.get("code") != 0:
        raise Exception(f"Lark token 失败: {resp.get('msg')}")
    return resp["tenant_access_token"]


def wiki_to_obj_token(token, wiki_token):
    """wiki node token → obj_token (spreadsheet token)"""
    url = f"https://open.larksuite.com/open-apis/wiki/v2/spaces/get_node?token={wiki_token}"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    resp = json.loads(urllib.request.urlopen(req).read())
    if resp.get("code") != 0:
        raise Exception(f"wiki get_node 失败: code={resp.get('code')} msg={resp.get('msg')}")
    node = resp.get("data", {}).get("node", {})
    obj_token = node.get("obj_token")
    obj_type = node.get("obj_type")
    if not obj_token:
        raise Exception(f"wiki node 无 obj_token: {node}")
    log(f"    wiki {wiki_token[:12]}... → obj_token={obj_token[:12]}... ({obj_type})")
    return obj_token


def export_spreadsheet(token, obj_token, name):
    """export_tasks: type=sheet → 整个 spreadsheet 一次导出 xlsx"""
    log(f"  创建导出任务: {name}")
    data = json.dumps({"file_extension": "xlsx", "token": obj_token, "type": "sheet"}).encode()
    req = urllib.request.Request(
        "https://open.larksuite.com/open-apis/drive/v1/export_tasks",
        data=data,
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
    )
    resp = json.loads(urllib.request.urlopen(req).read())
    if resp.get("code") != 0:
        raise Exception(f"创建导出任务失败: {resp}")
    ticket = resp["data"]["ticket"]

    for i in range(60):
        time.sleep(3)
        url = f"https://open.larksuite.com/open-apis/drive/v1/export_tasks/{ticket}?token={obj_token}"
        req2 = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
        resp2 = json.loads(urllib.request.urlopen(req2).read())
        result = resp2.get("data", {}).get("result", {})
        status = result.get("job_status")
        if status == 0:
            file_token = result["file_token"]
            file_size = result.get("file_size", 0)
            req3 = urllib.request.Request(
                f"https://open.larksuite.com/open-apis/drive/v1/export_tasks/file/{file_token}/download",
                headers={"Authorization": f"Bearer {token}"},
            )
            content = urllib.request.urlopen(req3).read()
            xlsx_path = os.path.join(EXPORT_DIR, f"_chat_{name}.xlsx")
            with open(xlsx_path, "wb") as f:
                f.write(content)
            log(f"  ✅ {name} xlsx {file_size/1024:.0f}KB → {xlsx_path}")
            return xlsx_path
        elif status == 2:
            raise Exception(f"导出任务失败: {resp2}")

    raise Exception("导出超时（180s）")


def parse_xlsx_write_csvs(xlsx_path, sheet_to_csv, name):
    """提取 xlsx 多 sheet → 写每个 csv"""
    try:
        import openpyxl
    except ImportError:
        os.system("pip3 install openpyxl -q")
        import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheets_in_file = list(wb.sheetnames)
    log(f"  {name} xlsx 含 sheet: {sheets_in_file}")

    written = 0
    for sheet_name, csv_file in sheet_to_csv.items():
        if sheet_name not in sheets_in_file:
            log(f"    ⚠️ sheet「{sheet_name}」在 xlsx 里找不到，跳过")
            continue
        ws = wb[sheet_name]
        csv_path = os.path.join(EXPORT_DIR, csv_file)
        rows_written = 0
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                vals = ["" if v is None else str(v) for v in row]
                w.writerow(vals)
                rows_written += 1
        log(f"    ✅ {csv_file}: {rows_written} 行")
        written += 1
    wb.close()
    return written


def main():
    config = json.load(open(CONFIG_PATH))
    if not config.get("lark"):
        log("❌ 无 Lark 配置")
        return 1

    log("=" * 60)
    log("Lark 聊天文档全量导出开始（API 方式，替代 Playwright UI）")

    token = get_lark_token(config)
    log(f"Lark token 获取成功")

    if not os.path.exists(EXPORT_DIR):
        os.makedirs(EXPORT_DIR)

    total_csvs = 0
    failed_docs = 0

    for doc in CHAT_DOCS:
        log(f"\n📄 {doc['name']} ({len(doc['sheet_to_csv'])} 个 sheet)")
        try:
            obj_token = wiki_to_obj_token(token, doc["wiki_token"])
            xlsx_path = export_spreadsheet(token, obj_token, doc["name"])
            n = parse_xlsx_write_csvs(xlsx_path, doc["sheet_to_csv"], doc["name"])
            total_csvs += n
            time.sleep(5)
        except Exception as e:
            log(f"  ❌ {doc['name']} 失败: {e}")
            failed_docs += 1

    log("\n" + "=" * 60)
    log(f"✅ 总计 {total_csvs} 个 csv 写入，失败文档 {failed_docs} 个")
    return 0 if failed_docs == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
